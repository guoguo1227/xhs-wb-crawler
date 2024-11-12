import requests
import openpyxl

from openpyxl import Workbook, load_workbook
from requests.packages.urllib3.exceptions import InsecureRequestWarning
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

class Home:
    def __init__(self):
        self.search_url = "https://api.bilibili.com/x/web-interface/wbi/search/type?category_id=&search_type=media_bangumi&ad_resource=5646&__refresh__=true&_extra=&context=&page=1&page_size=12&order=&pubtime_begin_s=0&pubtime_end_s=0&duration=&from_source=&from_spmid=333.337&platform=pc&highlight=1&single_column=0&keyword="
        self.detail_url = "https://api.bilibili.com/pgc/view/web/ep/list?season_id="
        #替换浏览器cookie
        self.cookie = ""
    def get_headers(self):
        return {
            "accept": "application/json",
            "accept-language": "zh-CN,zh;q=0.9",
            "content-type": "application/json;charset=UTF-8",
            "referer": "https://www.bilibili.com/bangumi/play/ss2727",
            "cookie": self.cookie,
            "user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36"
        }

    def detail(self, id, keyword, wb, ws):
        url = self.detail_url + str(id)
        response = requests.get(url, headers=self.get_headers(), verify=False)
        name = ''
        title = ''
        try:
            list = response.json()['result']['episodes']
            if len(list) != 0:
                name = list[0]['share_copy']
                title = list[0]['subtitle']
            else:
                print(f"search detail url:{url} no result")

            ws.append([keyword, name, title])
            wb.save('动画播放数列表.xlsx')
            print(f"搜索{keyword},完成")

        except Exception as e:
            print(f'搜索：{url},失败 {e}')

    def search_keyword(self, keyword):
        url = self.search_url + keyword
        response = requests.get(url, headers=self.get_headers(), verify=False)
        try:
            list = response.json()['data']['result']
            #获取第一个动画
            if len(list) != 0:
                return list[0]['season_id']
        except Exception as e:
            print(f'搜索：{keyword},失败 {e}')


    def main(self):
        # 创建一个新的工作簿
        wb = Workbook()
        # 选择默认的工作表
        ws = wb.active
        ws.append(['搜索词', '动画标题', '播放数'])

        # 打开 Excel 文件
        workbook = openpyxl.load_workbook('./input/二次元动画.xlsx')
        # 获取第一个工作表
        worksheet = workbook.active
        # 读取单元格内容
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            name = row[0]
            #搜索动画名称，拿到动画的ID
            season_id = self.search_keyword(name)
            if season_id:
                #获取动画播放量
                self.detail(season_id, name, wb, ws)

if __name__ == '__main__':
    home = Home()
    home.main()

#
# B站动画播放量爬虫
