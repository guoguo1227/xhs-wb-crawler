import requests
import openpyxl
from openpyxl import Workbook, load_workbook


class Home:
    def __init__(self, ):
        self.url = "https://app-ios-cdn.jjwxc.com/iosapi/novelbasicinfo?novelId="
        self.jump_url = "https://www.jjwxc.net/onebook.php?novelid="
        #替换浏览器cookie
        self.cookie = ""
    def get_headers(self):
        return {
            "accept": "application/json",
            "accept-language": "zh-CN,zh;q=0.9",
            "content-type": "application/json;charset=UTF-8",
            "cookie": self.cookie,
            "user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36"
        }

    def detail(self, note_id, keyword, wb, ws):
        url = self.url + str(note_id)
        response = requests.get(url, headers=self.get_headers())

        try:
            data = response.json()
            novelname = data['novelName']
            clicks = data['novip_clicks']
            jumpurl = self.jump_url + str(note_id)
            ws.append([keyword, novelname, jumpurl, clicks])
            wb.save('晋江文学城小说点击数列表.xlsx')

        except Exception as e:
            print(f'搜索：{url},失败 {e}')

    def search_keyword(self, keyword):
        url = f"http://ios.jjwxc.com/iosapi/search?filterFav=0&keyword={keyword}&page=1&searchType=8&sortMode=DESC&token=&type=1&versionCode=604"
        response = requests.get(url, headers=self.get_headers(), verify=False)
        try:
            data = response.json()['items']
            if len(data) != 0:
                return data[0]['novelid']
        except Exception as e:
            print(f'搜索：{keyword},失败 {e}')


    def main(self):
        # 创建一个新的工作簿
        wb = Workbook()
        # 选择默认的工作表
        ws = wb.active
        ws.append(['IP团体', '小说名称', '小说链接', '点击数（非V章均）'])

        # 打开 Excel 文件
        workbook = openpyxl.load_workbook('./input/二次元小说.xlsx')
        # 获取第一个工作表
        worksheet = workbook.active
        # 读取单元格内容
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            name = row[0]
            #根据关键词搜索小说ID
            note_id = self.search_keyword(name)
            if note_id:
                #获取小说点击数
                self.detail(note_id, name, wb, ws)


if __name__ == '__main__':
    home = Home()
    home.main()

